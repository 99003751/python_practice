def get_data(workbook, sheet_name, ps_number,):
    mastersheet = workbook.get_sheet_by_name("mastersheet")
    for r in range(1, sheet_name.max_row + 1):
        # to check the value in column 1
        if (sheet_name.cell(row=r, column=1).value == ps_number):
            # to traverse through the columns
            for c in range(2, sheet_name.max_column + 1):
                # to get all the values
                print(sheet_name.cell(row=1, column=c).value, sheet_name.cell(row=r, column=c).value)
                mastersheet.append([sheet_name.cell(row=1, column=c).value, sheet_name.cell(row=r, column=c).value])
                workbook.save("student.xlsx")

from openpyxl import load_workbook
workbook = load_workbook("student.xlsx")
sheet_list = workbook.get_sheet_names()
if 'mastersheet' in workbook.sheetnames:
    remove_mastersheet = workbook['mastersheet']
    workbook.remove(remove_mastersheet)
workbook.create_sheet("mastersheet")
ps_number = int(input("enter ps number"))
for i in range(0, len(sheet_list)):
    sheetName = workbook.get_sheet_by_name(sheet_list[i])
    get_data(workbook, sheetName, ps_number)


