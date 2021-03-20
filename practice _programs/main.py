# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

# type() for knowing the data type
#  id() for knowing the memory address

# importing libraries
# import pandas as pd
# # import openpyxl
# # import xlsxwriter
# # import xlrd
#
#
# all_df = pd.DataFrame()
#
# sheet0 = pd.read_excel(r'student.xlsx', sheet_name="Sheet1")
#
# all_df = all_df.append(sheet0)
# sheet1 = pd.read_excel(r'student.xlsx', sheet_name="Sheet2")
# sheet2 = pd.read_excel(r'student.xlsx', sheet_name="Sheet3")
# sheet3 = pd.read_excel(r'student.xlsx', sheet_name="Sheet4")
# sheet4 = pd.read_excel(r'student.xlsx', sheet_name="Sheet5")
#
# all_df = pd.merge(all_df, sheet1, how='left')
# # print(all_df)
# all_df = pd.merge(all_df, sheet2, how='left')
# # print(all_df)
# all_df = pd.merge(all_df, sheet3, how='left')
# # print(all_df)
# all_df = pd.merge(all_df, sheet4, how='left')
# print(all_df)
# df = pd.DataFrame(all_df, columns=["Name", "Roll no",
#                                    "physem1", "biosem1", "chemistrysem1", "mathssem1", "electronicssem1",
#                                    "phylabsem1", "biolabsem1", "chemlabsem1",
#
#                                     "physsem2", "biosem2", "chemistrysem2",
#                                    "mathssem2", "electronicssem2", "phylabsem2", "biolabsem2", "chemlabsem2",
#
#                                    "physem3","biosem3", "chemistrysem3", "mathssem3", "electronicssem3", "phylabsem3",
#                                    "biolabsem3", "chemlabsem3",
#
#                                    "physem4", "biosem4", "chemistrysem4", "mathssem4",
#                                    "electronicssem4", "phylabsem4", "biolabsem4", "chemlabsem4",
#
#                                    "physem5", "biosem5",
#                                    "chemistrysem5", "mathssem5", "electronicssem5", "phylabsem5", "biolabsem5",
#                                    "chemlabsem5"])
#
# i = int(input("enter roll number"))
# dfmaster = pd.DataFrame()
# print(df)
# # if len(df)==0:
# #    print("empty data frame")
# # else:
# #     print("data is available")
#
# dfmaster = df.loc[df["Roll no"] == i]
# if len(dfmaster)==0:
#    print("empty data frame")
# else:
#     print("data is available")
#
#
#
# # name=input('enter student name to fetch the data')
# #
# # # df.set_index("NAME", inplace=True)
# #
# # dfmaster1=df.loc[df["NAME"]==name]
# #
# # if len(dfmaster1)==0:
# #    print("empty data frame")
# # else:
# #     print("non empty data frame")
#
#
# from openpyxl import load_workbook
#
# path = r"student.xlsx"
#
# book = load_workbook(path)
# writer = pd.ExcelWriter(path, engine='openpyxl')
# writer.book = book
# shlist=book.sheetnames
# # print(shlist)
# if 'mastersheet' in book.sheetnames:
#    # msheet=book.get_sheet_by_name('mastersheet')
#     jkl = book['mastersheet']
#     book.remove(jkl)
#
#    # dfmaster.to_excel(writer, sheet_name='mastersheet')
#
# dfmaster.to_excel(writer, sheet_name='mastersheet')
#
# writer.save()
# writer.close()
def print_by_sheet(sheet_name,ps_name):
    for r in range(1, sheet_name.max_row + 1):
        # to check the value in column 1
        if (sh.cell(row=r, column=1).value == 2):
            # to traverse through the columns
            for c in range(2, sheet_name.max_column + 1):
                # to get all the values
                print(sh.cell(row=1, column=c).value, sh.cell(row=r, column=c).value)

from openpyxl import load_workbook
# load excel with its path
wrkbk = load_workbook("student.xlsx")
# to get the active work sheet
sh = wrkbk.get_sheet_by_name("Sheet1")
ps_name=int(input("enter ps number"))
print_by_sheet(sh, ps_name)

# to print the maximum number of occupied rows in console
print(sh.max_row)
# to print the maximum number of occupied columns in console
print(sh.max_column)
# to get all the values from the excel and traverse through the rows

#
# def print_by_sheet(sheet_name):
#     for r in range(1, sheet_name.max_row + 1):
#         # to check the value in column 1
#         if (sh.cell(row=r, column=1).value == 2):
#             # to traverse through the columns
#             for c in range(2, sheet_name.max_column + 1):
#                 # to get all the values
#                 print(sh.cell(row=1, column=c).value, sh.cell(row=r, column=c).value)

