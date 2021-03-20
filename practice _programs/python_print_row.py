# class one:
#     print("hello this is class")
#     def __init__(self,name):
#         self.clsname=name
#     def print_cls_const(self):
#         print(self.clsname)
#     def print_cls_one(self,name):
#         print("hello "+str(name))
#
# x=one(input("enter name to pass in the cls "))
# x.print_cls_one("shrinidhi")
# x.print_cls_const()


class Data:
    count=3
    def get_workbook(self, filename):
        return load_workbook(filename)

    def get_sheet_list(self, workbook):
        return workbook.get_sheet_names()

    def create_mastersheet(self, workbook, mastersheet):
        if mastersheet in self.get_sheet_list(workbook):
            # msheet = workbook.get_sheet_by_name(mastersheet)
            jkl = workbook[mastersheet]
            workbook.remove(jkl)
        workbook.create_sheet("mastersheet")

    def get_user_input(self):
        print("choose option\n1.PS number\n2.Name\n3.Email id")
        a = int(input())
        return a

    def get_data(self, workbook, sheet_name, ps_number, a):
        mastersheet = workbook.get_sheet_by_name("mastersheet")
        for r in range(1, sheet_name.max_row + 1):
            # to check the value in column 1
            if sheet_name.cell(row=r, column=a).value == ps_number:
                global flag
                if flag == True:
                    flag = False
                    mastersheet.cell(row=1, column=1).value = sheet_name.cell(row=1, column=1).value
                    mastersheet.cell(row=1, column=2).value = sheet_name.cell(row=1, column=2).value
                    mastersheet.cell(row=1, column=3).value = sheet_name.cell(row=1, column=3).value

                    if a == 1:
                        mastersheet.cell(row=2, column=1).value = sheet_name.cell(row=r, column=a).value
                        mastersheet.cell(row=2, column=2).value = sheet_name.cell(row=r, column=a+1).value
                        mastersheet.cell(row=2, column=3).value = sheet_name.cell(row=r, column=a+2).value
                    elif a == 2:
                        mastersheet.cell(row=2, column=1).value = sheet_name.cell(row=r, column=a-1).value
                        mastersheet.cell(row=2, column=2).value = sheet_name.cell(row=r, column=a).value
                        mastersheet.cell(row=2, column=3).value = sheet_name.cell(row=r, column=a +1).value
                    elif a == 3:
                        mastersheet.cell(row=2, column=1).value = sheet_name.cell(row=r, column=a-2).value
                        mastersheet.cell(row=2, column=2).value = sheet_name.cell(row=r, column=a-1).value
                        mastersheet.cell(row=2, column=3).value = sheet_name.cell(row=r, column=a).value
                # to traverse through the columns
                for c in range(4, sheet_name.max_column + 1):
                    # to get all the values
                    # print(sheet_name.cell(row=1, column=c).value, sheet_name.cell(row=r, column=c).value)
                    # mastersheet.append([sheet_name.cell(row=1, column=c).value, sheet_name.cell(row=r, column=c).value])
                    mastersheet.cell(row=1, column=Data.count+1).value = sheet_name.cell(row=1, column=c).value
                    mastersheet.cell(row=2, column=Data.count+1).value = sheet_name.cell(row=r, column=c).value
                    workbook.save("student.xlsx")


from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# workbook = load_workbook("student.xlsx")
# sheet_list=workbook.get_sheet_names()
# if 'mastersheet' in workbook.sheetnames:
#     msheet=workbook.get_sheet_by_name('mastersheet')
#     jkl = workbook['mastersheet']
#     workbook.remove(jkl)
# workbook.create_sheet("mastersheet")
#
# print("choose option\n1.PS number\n2.Name\n3.Email id")
# a=int(input())
flag = True
obj = Data()
workbook1 = obj.get_workbook("student.xlsx")
sheet_list = obj.get_sheet_list(workbook1)
obj.create_mastersheet(workbook1, "mastersheet")
user_input = obj.get_user_input()
if user_input == 1:
    ps_number = int(input("enter ps number"))
    for i in range(0, len(sheet_list)):
        sheetName = workbook1.get_sheet_by_name(sheet_list[i])
        obj.get_data(workbook1, sheetName, ps_number, user_input)
elif user_input == 2:
    name = input("enter name")
    for i in range(0, len(sheet_list)):
        sheetName = workbook1.get_sheet_by_name(sheet_list[i])
        obj.get_data(workbook1, sheetName, name, user_input)
elif user_input == 3:
    email_id = input("Enter email id")
    for i in range(0, len(sheet_list)):
        sheetName = workbook1.get_sheet_by_name(sheet_list[i])
        obj.get_data(workbook1, sheetName, email_id, user_input)
else:
    print("Invalid input\n")

mastersheet1 = workbook1.get_sheet_by_name("mastersheet")
values = Reference(mastersheet1, min_col=1, min_row=4,
                   max_col=2, max_row=38)

chart = BarChart()
chart.add_data(values)
chart.title = " BAR-CHART "
chart.x_axis.title = " X_AXIS "
chart.y_axis.title = " Y_AXIS "
mastersheet1.add_chart(chart, "E2")
workbook1.save("student.xlsx")


#
# mastersheet.cell(row=1, column=c).value = sheet_name.cell(row=1, column=c).value
# mastersheet.cell(row=2, column=c).value = sheet_name.cell(row=r, column=c).value