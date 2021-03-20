class Data:
    count = 3

    def get_workbook(self, filename):
        return load_workbook(filename)

    def get_sheet_list(self, workbook):
        return workbook.sheetnames

    def create_mastersheet(self, workbook, mastersheet):
        if mastersheet in self.get_sheet_list(workbook):
            jkl = workbook[mastersheet]
            workbook.remove(jkl)
        workbook.create_sheet("mastersheet")

    def get_user_input(self):
        print("choose the search parameter\n1.PS number\n2.Name\n3.Email id")
        a = int(input())
        return a

    def get_data(self, workbook, sheet_name, ps_number, a, i):
        mastersheet = workbook.get_sheet_by_name("mastersheet")
        for r in range(1, sheet_name.max_row + 1):
            # to check the value in column 1
            if sheet_name.cell(row=r, column=a).value == ps_number:
                global flag
                if flag == True:
                    flag = False
                    mastersheet.cell(row=1, column=1).value = sheet_name.cell(row=1, column=1).value
                    mastersheet.cell(row=1, column=2).value = sheet_name.cell(row=1, column=2).value
                    mastersheet.cell(row=1, column=3).value = sheet_name.cell(row=1, column=3).value

                if a == 1:
                    mastersheet.cell(row=i + 2, column=1).value = sheet_name.cell(row=r, column=a).value
                    mastersheet.cell(row=i + 2, column=2).value = sheet_name.cell(row=r, column=a + 1).value
                    mastersheet.cell(row=i + 2, column=3).value = sheet_name.cell(row=r, column=a + 2).value
                elif a == 2:
                    mastersheet.cell(row=i + 2, column=1).value = sheet_name.cell(row=r, column=a - 1).value
                    mastersheet.cell(row=i + 2, column=2).value = sheet_name.cell(row=r, column=a).value
                    mastersheet.cell(row=i + 2, column=3).value = sheet_name.cell(row=r, column=a + 1).value
                elif a == 3:
                    mastersheet.cell(row=i + 2, column=1).value = sheet_name.cell(row=r, column=a - 2).value
                    mastersheet.cell(row=i + 2, column=2).value = sheet_name.cell(row=r, column=a - 1).value
                    mastersheet.cell(row=i + 2, column=3).value = sheet_name.cell(row=r, column=a).value
                # to traverse through the columns
                for c in range(4, sheet_name.max_column + 1):
                    # to get all the values
                    mastersheet.cell(row=1, column=Data.count + 1).value = sheet_name.cell(row=1, column=c).value
                    mastersheet.cell(row=i + 2, column=Data.count + 1).value = sheet_name.cell(row=r, column=c).value
                    Data.count = Data.count + 1
                    workbook.save("student.xlsx")


def do_function(j):
    Data.count = 3
    if user_input == 1:
        ps_number = int(input("enter ps number"))
        for i in range(0, len(sheet_list) - 1):
            sheetName = workbook1.get_sheet_by_name(sheet_list[i])
            data_object.get_data(workbook1, sheetName, ps_number, user_input, j)
    elif user_input == 2:
        name = input("enter name")
        for i in range(0, len(sheet_list) - 1):
            sheetName = workbook1.get_sheet_by_name(sheet_list[i])
            data_object.get_data(workbook1, sheetName, name, user_input, j)
    elif user_input == 3:
        email_id = input("Enter email id")
        for i in range(0, len(sheet_list) - 1):
            sheetName = workbook1.get_sheet_by_name(sheet_list[i])
            data_object.get_data(workbook1, sheetName, email_id, user_input, j)
    else:
        print("Invalid input\n")


from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

flag = True
data_object = Data()
workbook1 = data_object.get_workbook("student.xlsx")
sheet_list = data_object.get_sheet_list(workbook1)
data_object.create_mastersheet(workbook1, "mastersheet")
n = int(input("How many students data you want...?"))
for i in range(n):
    user_input = data_object.get_user_input()
    do_function(i)

mastersheet1 = workbook1.get_sheet_by_name("mastersheet")
values = Reference(mastersheet1, min_col=4, min_row=2, max_col=38, max_row=n + 1)
chart = BarChart()
chart.add_data(values)
chart.title = " BAR-CHART "
chart.x_axis.title = " X_AXIS "
chart.y_axis.title = " Y_AXIS "
mastersheet1.add_chart(chart, "E2")
workbook1.save("student.xlsx")
