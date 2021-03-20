flag = 10

def get_data(workbook, sheet_name, ps_number, a):
    mastersheet = workbook.get_sheet_by_name("mastersheet")
    for r in range(1, sheet_name.max_row + 1):
        # to check the value in column 1
        if (sheet_name.cell(row=r, column=a).value == ps_number):
            global flag
            if( flag== 10):
                flag = 11
                if(a == 1):
                    mastersheet.append([sheet_name.cell(row=1, column=1).value, sheet_name.cell(row=r, column=a).value])
                    mastersheet.append([sheet_name.cell(row=1, column=2).value, sheet_name.cell(row=r, column=a + 1).value])
                    mastersheet.append([sheet_name.cell(row=1, column=3).value, sheet_name.cell(row=r, column=a + 2).value])
                    workbook.save("student.xlsx")
                elif(a == 2):
                    mastersheet.append([sheet_name.cell(row=1, column=1).value, sheet_name.cell(row=r, column=a-1).value])
                    mastersheet.append([sheet_name.cell(row=1, column=2).value, sheet_name.cell(row=r, column=a).value])
                    mastersheet.append([sheet_name.cell(row=1, column=3).value, sheet_name.cell(row=r, column=a+1).value])
                    workbook.save("student.xlsx")
                elif (a == 3):
                    mastersheet.append([sheet_name.cell(row=1, column=1).value, sheet_name.cell(row=r, column=a - 2).value])
                    mastersheet.append([sheet_name.cell(row=1, column=2).value, sheet_name.cell(row=r, column=a-1).value])
                    mastersheet.append([sheet_name.cell(row=1, column=3).value, sheet_name.cell(row=r, column=a).value])
                    workbook.save("student.xlsx")
            # to traverse through the columns
            for c in range(4, sheet_name.max_column + 1):
                # to get all the values
                # print(sheet_name.cell(row=1, column=c).value, sheet_name.cell(row=r, column=c).value)
                mastersheet.append([sheet_name.cell(row=1, column=c).value, sheet_name.cell(row=r, column=c).value])
                workbook.save("student.xlsx")


from openpyxl import load_workbook
from openpyxl.chart import BarChart,Reference
workbook = load_workbook("student.xlsx")
sheet_list=workbook.get_sheet_names()
if 'mastersheet' in workbook.sheetnames:
    msheet=workbook.get_sheet_by_name('mastersheet')
    jkl = workbook['mastersheet']
    workbook.remove(jkl)
workbook.create_sheet("mastersheet")

print("choose option\n1.PS number\n2.Name\n3.Email id")
a=int(input())
if(a==1):
    ps_number = int(input("enter ps number"))
    for i in range(0, len(sheet_list)):
        sheetName = workbook.get_sheet_by_name(sheet_list[i])
        get_data(workbook, sheetName, ps_number, a)
elif(a==2):
    name=input("enter name")
    for i in range(0, len(sheet_list)):
        sheetName = workbook.get_sheet_by_name(sheet_list[i])
        get_data(workbook, sheetName, name, a)
elif(a==3):
    email_id=input("Enter email id")
    for i in range(0, len(sheet_list)):
        sheetName = workbook.get_sheet_by_name(sheet_list[i])
        get_data(workbook, sheetName, email_id, a)
else:
    print("Invalid input\n")

mastersheet = workbook.get_sheet_by_name("mastersheet")


values = Reference(mastersheet, min_col = 1, min_row = 4,
                         max_col = 2, max_row = 38)

chart = BarChart()
chart.add_data(values)
chart.title = " BAR-CHART "
chart.x_axis.title = " X_AXIS "
chart.y_axis.title = " Y_AXIS "
mastersheet.add_chart(chart, "E2")
workbook.save("student.xlsx")